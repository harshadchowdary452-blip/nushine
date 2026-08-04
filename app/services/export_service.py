import csv
import io
import os
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from fastapi.responses import StreamingResponse
from app.config import settings
from app.models.patient import Patient
from app.models.appointment import Appointment
from app.models.case import Case
from app.models.treatment_plan import TreatmentPlan
from app.models.billing import Billing
from app.models.hospital_monthly_expense import HospitalMonthlyExpense
from app.models.lead import Lead
from app.models.enquiry import Enquiry
from app.models.follow_up import FollowUp
from app.models.user import User
from app.models.consent_form import ConsentForm
from app.models.hospital import Hospital
from app.models.audit_log import AuditLog
from app.utils.dashboard_helpers import get_date_range, calculate_expenses_for_date_range

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

# ── Scope helpers ──────────────────────────────────────────────────

async def _get_hospital_ids(db: AsyncSession, current_user: dict):
    role = current_user.get("role")
    if role == "SUPER_ADMIN":
        return None
    if role == "GROUP_ADMIN":
        agid = current_user.get("admin_group_id")
        r = await db.execute(select(Hospital.id).where(Hospital.admin_group_id == agid))
        return [row[0] for row in r.all()]
    hid = current_user.get("hospital_id")
    return [hid] if hid else None


async def _get_patient_ids(db, hospital_ids):
    if not hospital_ids:
        return None
    r = await db.execute(select(Patient.id).where(Patient.hospital_id.in_(hospital_ids)))
    return [row[0] for row in r.all()]


async def _get_case_ids(db, patient_ids):
    if not patient_ids:
        return None
    r = await db.execute(select(Case.id).where(Case.patient_id.in_(patient_ids)))
    return [row[0] for row in r.all()]


async def _name_dict(db, model, ids):
    """Return {id: full_name} for a set of user/patient IDs."""
    if not ids:
        return {}
    r = await db.execute(select(model.id, model.full_name).where(model.id.in_(ids)))
    return {row[0]: row[1] for row in r.all()}


# ── Hospital info for PDF headers ────────────────────────────────────

async def _hospital_info(db, hospital_id: str = None):
    """Return name, address, phone for a hospital."""
    if not hospital_id:
        return {"name": "Hospital", "address": "", "phone": ""}
    r = await db.execute(
        select(Hospital.name, Hospital.address, Hospital.phone).where(Hospital.id == hospital_id)
    )
    row = r.first()
    if row:
        row = row._mapping  # avoid __getitem__ issues in async
        return {"name": row["name"] or "Hospital", "address": row["address"] or "", "phone": row["phone"] or ""}
    return {"name": "Hospital", "address": "", "phone": ""}


# ═══════════════════════════════════════════════════════════════════
#  CSV GENERATORS  –  business columns only, no IDs/timestamps
# ═══════════════════════════════════════════════════════════════════
import sys
sys.setrecursionlimit(10000)


def _stream_csv(rows, headers, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


async def _csv_patients(db, hospital_ids, date_start, date_end):
    q = select(
        Patient.op_no, Patient.full_name, Patient.age, Patient.gender,
        Patient.phone, Patient.abha_id, Patient.patient_source, Patient.doctor_id, Patient.status,
    ).where(Patient.created_at >= date_start, Patient.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(Patient.hospital_id.in_(hospital_ids))
    q = q.order_by(Patient.created_at.desc())
    rows = (await db.execute(q)).all()
    doc_ids = set(r[7] for r in rows if r[7])
    doc_names = await _name_dict(db, User, doc_ids)
    headers = ["OP Number", "Patient Name", "Age", "Gender", "Phone", "ABHA ID", "Source", "Assigned Doctor", "Status"]
    data = [[r[0] or "", r[1], r[2] or "", r[3] or "", r[4] or "", r[5] or "", r[6] or "", doc_names.get(r[7], ""), r[8]] for r in rows]
    return data, headers, len(rows)


async def _csv_appointments(db, hospital_ids, date_start, date_end):
    pids = await _get_patient_ids(db, hospital_ids)
    q = select(
        Appointment.appointment_number, Appointment.patient_id, Appointment.doctor_id,
        Appointment.appointment_date, Appointment.appointment_time, Appointment.status,
    ).where(Appointment.appointment_date >= date_start.date() if hasattr(date_start, 'date') else date_start,
             Appointment.appointment_date < date_end.date() if hasattr(date_end, 'date') else date_end)
    if pids is not None:
        q = q.where(Appointment.patient_id.in_(pids))
    q = q.order_by(Appointment.appointment_date.desc())
    rows = (await db.execute(q)).all()
    pids = set(r[1] for r in rows if r[1])
    dids = set(r[2] for r in rows if r[2])
    pnames = await _name_dict(db, Patient, pids)
    dnames = await _name_dict(db, User, dids)
    headers = ["Appointment Number", "Patient Name", "Doctor Name", "Date", "Time", "Status"]
    data = [[r[0], pnames.get(r[1], ""), dnames.get(r[2], ""), str(r[3]) if r[3] else "", str(r[4]) if r[4] else "", r[5]] for r in rows]
    return data, headers, len(rows)


async def _csv_cases(db, hospital_ids, date_start, date_end):
    pids = await _get_patient_ids(db, hospital_ids)
    q = select(
        Case.case_number, Case.patient_id, Case.doctor_id,
        Case.chief_complaint, Case.diagnosis, Case.status,
    ).where(Case.created_at >= date_start, Case.created_at < date_end)
    if pids is not None:
        q = q.where(Case.patient_id.in_(pids))
    q = q.order_by(Case.created_at.desc())
    rows = (await db.execute(q)).all()
    pat_ids = set(r[1] for r in rows if r[1])
    doc_ids = set(r[2] for r in rows if r[2])
    pnames = await _name_dict(db, Patient, pat_ids)
    dnames = await _name_dict(db, User, doc_ids)
    headers = ["Case Number", "Patient Name", "Doctor Name", "Chief Complaint", "Diagnosis", "Status"]
    data = [[r[0], pnames.get(r[1], ""), dnames.get(r[2], ""), r[3] or "", r[4] or "", r[5]] for r in rows]
    return data, headers, len(rows)


async def _csv_treatments(db, hospital_ids, date_start, date_end):
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)
    q = select(
        TreatmentPlan.id, TreatmentPlan.treatment_name, TreatmentPlan.case_id,
        TreatmentPlan.total_sittings, TreatmentPlan.completed_sittings,
        TreatmentPlan.remaining_sittings, TreatmentPlan.status,
    ).where(TreatmentPlan.created_at >= date_start, TreatmentPlan.created_at < date_end)
    if cids is not None:
        q = q.where(TreatmentPlan.case_id.in_(cids))
    q = q.order_by(TreatmentPlan.created_at.desc())
    rows = (await db.execute(q)).all()
    case_ids = set(r[2] for r in rows if r[2])
    case_map = {}
    if case_ids:
        cr = await db.execute(select(Case.id, Case.patient_id, Case.doctor_id).where(Case.id.in_(case_ids)))
        for case_row in cr.all():
            case_map[case_row[0]] = (case_row[1], case_row[2])
    pat_ids = set(v[0] for v in case_map.values() if v[0])
    doc_ids = set(v[1] for v in case_map.values() if v[1])
    pnames = await _name_dict(db, Patient, pat_ids)
    dnames = await _name_dict(db, User, doc_ids)
    headers = ["Treatment Name", "Patient Name", "Doctor Name", "Total Sittings", "Completed Sittings", "Remaining Sittings", "Progress %", "Status"]
    data = []
    for r in rows:
        ts, cs, rs = r[3] or 1, r[4] or 0, r[5] or 1
        pct = round((cs / ts) * 100, 1) if ts > 0 else 0
        pat_name = ""
        doc_name = ""
        if r[2] in case_map:
            pid, did = case_map[r[2]]
            pat_name = pnames.get(pid, "")
            doc_name = dnames.get(did, "")
        data.append([r[1] or "", pat_name, doc_name, ts, cs, rs, f"{pct}%", r[6]])
    return data, headers, len(rows)


async def _csv_billings(db, hospital_ids, date_start, date_end):
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)
    q = select(
        Billing.invoice_number, Billing.case_id, Billing.total_amount,
        Billing.paid_amount, Billing.pending_amount, Billing.payment_status,
        Billing.created_at,
    ).where(Billing.created_at >= date_start, Billing.created_at < date_end)
    if cids is not None:
        q = q.where(Billing.case_id.in_(cids))
    q = q.order_by(Billing.created_at.desc())
    rows = (await db.execute(q)).all()
    case_ids = set(r[1] for r in rows if r[1])
    case_map = {}
    if case_ids:
        cr = await db.execute(
            select(Case.id, Case.patient_id, Case.doctor_id).where(Case.id.in_(case_ids))
        )
        for cr_row in cr.all():
            case_map[cr_row[0]] = (cr_row[1], cr_row[2])
    pat_ids = set(v[0] for v in case_map.values() if v[0])
    doc_ids = set(v[1] for v in case_map.values() if v[1])
    pnames = await _name_dict(db, Patient, pat_ids)
    dnames = await _name_dict(db, User, doc_ids)
    op_map = {}
    if pat_ids:
        op_r = await db.execute(select(Patient.id, Patient.op_no).where(Patient.id.in_(pat_ids)))
        for op_row in op_r.all():
            op_map[op_row[0]] = op_row[1] or ""
    headers = ["Invoice Number", "Patient Name", "OP Number", "Doctor Name", "Bill Amount", "Paid Amount", "Pending Amount", "Payment Status", "Invoice Date"]
    data = []
    for r in rows:
        pat_name = ""
        doc_name = ""
        op_no = ""
        if r[1] in case_map:
            pid, did = case_map[r[1]]
            pat_name = pnames.get(pid, "")
            doc_name = dnames.get(did, "")
            op_no = op_map.get(pid, "")
        data.append([r[0] or "", pat_name, op_no, doc_name, float(r[2]), float(r[3]), float(r[4]), r[5], str(r[6].date()) if r[6] else ""])
    return data, headers, len(rows)


async def _csv_expenses(db, hospital_ids, date_start, date_end):
    sd = date_start.date() if hasattr(date_start, 'date') else date_start
    ed = date_end.date() if hasattr(date_end, 'date') else date_end
    q = select(
        HospitalMonthlyExpense.expense_date, HospitalMonthlyExpense.expense_category,
        HospitalMonthlyExpense.expense_name, HospitalMonthlyExpense.amount,
        HospitalMonthlyExpense.payment_method, HospitalMonthlyExpense.vendor,
        HospitalMonthlyExpense.created_by,
    ).where(HospitalMonthlyExpense.expense_date >= sd, HospitalMonthlyExpense.expense_date < ed)
    if hospital_ids is not None:
        q = q.where(HospitalMonthlyExpense.hospital_id.in_(hospital_ids))
    q = q.order_by(HospitalMonthlyExpense.expense_date.desc())
    rows = (await db.execute(q)).all()
    uids = set(r[6] for r in rows if r[6])
    unames = await _name_dict(db, User, uids)
    headers = ["Expense Date", "Category", "Expense Name", "Amount", "Payment Method", "Vendor", "Created By"]
    data = []
    for r in rows:
        data.append([
            str(r[0]) if r[0] else "",
            r[1] or "",
            r[2] or "",
            float(r[3] or 0),
            r[4] or "",
            r[5] or "",
            unames.get(r[6], ""),
        ])
    return data, headers, len(rows)


async def _csv_leads(db, hospital_ids, date_start, date_end):
    q = select(
        Lead.lead_name, Lead.mobile, Lead.email, Lead.source,
        Lead.interested_treatment, Lead.status, Lead.priority, Lead.lead_score,
        Lead.assigned_staff_id,
    ).where(Lead.created_at >= date_start, Lead.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(Lead.hospital_id.in_(hospital_ids))
    q = q.order_by(Lead.created_at.desc())
    rows = (await db.execute(q)).all()
    uids = set(r[8] for r in rows if r[8])
    unames = await _name_dict(db, User, uids)
    headers = ["Lead Name", "Mobile", "Email", "Source", "Interested Treatment", "Status", "Priority", "Score", "Assigned Staff"]
    data = [[r[0], r[1], r[2] or "", r[3] or "", r[4] or "", r[5], r[6], r[7], unames.get(r[8], "")] for r in rows]
    return data, headers, len(rows)


async def _csv_enquiries(db, hospital_ids, date_start, date_end):
    q = select(
        Enquiry.patient_id, Enquiry.treatment_interest,
        Enquiry.notes, Enquiry.status, Enquiry.assigned_staff_id,
        Enquiry.next_follow_up_date,
    ).where(Enquiry.created_at >= date_start, Enquiry.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(Enquiry.hospital_id.in_(hospital_ids))
    q = q.order_by(Enquiry.created_at.desc())
    rows = (await db.execute(q)).all()
    pids = set(r[0] for r in rows if r[0])
    uids = set(r[4] for r in rows if r[4])
    pnames = await _name_dict(db, Patient, pids)
    unames = await _name_dict(db, User, uids)
    op_map = {}
    if pids:
        op_r = await db.execute(select(Patient.id, Patient.op_no).where(Patient.id.in_(pids)))
        for op_row in op_r.all():
            op_map[op_row[0]] = op_row[1] or ""
    headers = ["Patient Name", "OP Number", "Treatment Interest", "Notes", "Status", "Assigned Staff", "Next Follow-up"]
    data = [[pnames.get(r[0], ""), op_map.get(r[0], ""), r[1] or "", r[2] or "", r[3], unames.get(r[4], ""), str(r[5]) if r[5] else ""] for r in rows]
    return data, headers, len(rows)


async def _csv_follow_ups(db, hospital_ids, date_start, date_end):
    q = select(
        FollowUp.patient_id, FollowUp.doctor_id, FollowUp.follow_up_date,
        FollowUp.follow_up_type, FollowUp.outcome, FollowUp.status,
        FollowUp.treatment_id,
    ).where(FollowUp.created_at >= date_start, FollowUp.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(FollowUp.hospital_id.in_(hospital_ids))
    q = q.order_by(FollowUp.created_at.desc())
    rows = (await db.execute(q)).all()
    pids = set(r[0] for r in rows if r[0])
    dids = set(r[1] for r in rows if r[1])
    pnames = await _name_dict(db, Patient, pids)
    dnames = await _name_dict(db, User, dids)
    op_map = {}
    if pids:
        op_r = await db.execute(select(Patient.id, Patient.op_no).where(Patient.id.in_(pids)))
        for op_row in op_r.all():
            op_map[op_row[0]] = op_row[1] or ""
    headers = ["Patient Name", "OP Number", "Doctor Name", "Follow-Up Type", "Due Date", "Status", "Outcome"]
    data = [[pnames.get(r[0], ""), op_map.get(r[0], ""), dnames.get(r[1], ""), r[3] or "", str(r[2]) if r[2] else "", r[5], r[4] or ""] for r in rows]
    return data, headers, len(rows)


async def _csv_recalls(db, hospital_ids, date_start, date_end):
    recall_types = ("6_MONTH_RECALL", "12_MONTH_RECALL", "CUSTOM_RECALL", "CUSTOM_FOLLOW_UP")
    q = select(
        FollowUp.patient_id, FollowUp.doctor_id, FollowUp.follow_up_date,
        FollowUp.follow_up_type, FollowUp.outcome, FollowUp.status,
        FollowUp.treatment_id,
    ).where(FollowUp.follow_up_type.in_(recall_types),
            FollowUp.created_at >= date_start, FollowUp.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(FollowUp.hospital_id.in_(hospital_ids))
    q = q.order_by(FollowUp.created_at.desc())
    rows = (await db.execute(q)).all()
    pids = set(r[0] for r in rows if r[0])
    dids = set(r[1] for r in rows if r[1])
    pnames = await _name_dict(db, Patient, pids)
    dnames = await _name_dict(db, User, dids)
    op_map = {}
    if pids:
        op_r = await db.execute(select(Patient.id, Patient.op_no).where(Patient.id.in_(pids)))
        for op_row in op_r.all():
            op_map[op_row[0]] = op_row[1] or ""
    headers = ["Patient Name", "OP Number", "Doctor Name", "Recall Type", "Recall Date", "Status"]
    rtype_labels = {
        "6_MONTH_RECALL": "6-Month Recall", "12_MONTH_RECALL": "12-Month Recall",
        "CUSTOM_RECALL": "Custom Follow-Up",
        "CUSTOM_FOLLOW_UP": "Custom Follow-Up",
    }
    data = [[pnames.get(r[0], ""), op_map.get(r[0], ""), dnames.get(r[1], ""), rtype_labels.get(r[3], r[3] or ""), str(r[2]) if r[2] else "", r[5]] for r in rows]
    return data, headers, len(rows)


async def _csv_doctors(db, hospital_ids, date_start, date_end):
    q = select(
        User.full_name, User.email, User.phone,
        User.specialization, User.is_active,
    ).where(User.role == "DOCTOR", User.created_at >= date_start, User.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(User.hospital_id.in_(hospital_ids))
    q = q.order_by(User.created_at.desc())
    rows = (await db.execute(q)).all()
    headers = ["Doctor Name", "Email", "Phone", "Specialization", "Active"]
    data = [[r[0], r[1] or "", r[2] or "", r[3] or "", "Yes" if r[4] else "No"] for r in rows]
    return data, headers, len(rows)


async def _csv_doctor_performance(db, current_user, date_start, date_end):
    """Performance & clinical productivity summary for the caller's scope."""
    from app.routers.doctor_performance import _collect, _doctors_in_scope, _hospital_scope_for_user

    doctors = await _doctors_in_scope(db, current_user, None, None)
    dids = [d.id for d in doctors]
    hospital_scope = _hospital_scope_for_user(current_user)
    metrics, _ = await _collect(db, dids, date_start, date_end, hospital_scope)

    headers = [
        "Doctor Name", "Designation", "Department", "Hospital ID",
        "Patients Seen", "New Patients", "Returning Patients",
        "Appointments Completed", "Appointments Cancelled",
        "Cases Created", "Cases Completed", "Active Cases",
        "Treatments Completed", "Sittings Completed",
        "Revenue", "Avg Revenue / Patient",
        "Attendance Rate", "Retention Rate", "Case Completion",
        "Treatment Acceptance", "Recall Success", "Avg Rating",
        "No Shows", "Outstanding Amount", "Cases With Reports",
    ]
    data = []
    for doc in doctors:
        m = metrics.get(doc.id, {})
        new_patients = max(0, m.get("patients_seen", 0) - m.get("returning_patients", 0))
        attendance_den = (m.get("appointments_completed", 0) +
                          m.get("appointments_cancelled", 0) +
                          m.get("appointments_rescheduled", 0))
        retention_den = new_patients + m.get("returning_patients", 0)
        recall_den = m.get("followups_completed", 0) + m.get("followups_lost", 0)
        from app.routers.doctor_performance import _pct

        data.append([
            doc.full_name,
            doc.qualification or "Doctor",
            doc.specialization or "General Dentistry",
            doc.hospital_id or "",
            m.get("patients_seen", 0),
            new_patients,
            m.get("returning_patients", 0),
            m.get("appointments_completed", 0),
            m.get("appointments_cancelled", 0),
            m.get("cases_created", 0),
            m.get("cases_completed_period", 0),
            m.get("active_cases", 0),
            m.get("treatments_completed", 0),
            m.get("sittings_completed", 0),
            round(m.get("revenue", 0), 2),
            round(m.get("revenue", 0) / m["patients_seen"], 2) if m.get("patients_seen") else 0.0,
            _pct(m.get("appointments_completed", 0), attendance_den),
            _pct(m.get("returning_patients", 0), retention_den),
            _pct(m.get("cases_completed_period", 0), m.get("cases_created", 0)),
            _pct(m.get("plans_created", 0), m.get("cases_created", 0)),
            _pct(m.get("followups_completed", 0), recall_den),
            round(m.get("rating_sum", 0) / m["rating_count"], 2) if m.get("rating_count") else "",
            m.get("no_shows", 0),
            round(m.get("outstanding_amount", 0), 2),
            m.get("cases_with_reports", 0),
        ])
    data.sort(key=lambda r: (r[14] if isinstance(r[14], (int, float)) else 0), reverse=True)
    return data, headers, len(data)


async def _csv_consent_forms(db, hospital_ids, date_start, date_end):
    q = select(
        ConsentForm.patient_id, ConsentForm.consent_type,
        ConsentForm.patient_name, ConsentForm.op_number, ConsentForm.phone,
        ConsentForm.doctor_id,
    ).where(ConsentForm.created_at >= date_start, ConsentForm.created_at < date_end,
            ConsentForm.is_deleted == False)
    if hospital_ids is not None:
        q = q.where(ConsentForm.hospital_id.in_(hospital_ids))
    q = q.order_by(ConsentForm.created_at.desc())
    rows = (await db.execute(q)).all()
    dids = set(r[5] for r in rows if r[5])
    dnames = await _name_dict(db, User, dids)
    headers = ["Patient Name", "Consent Type", "OP Number", "Phone", "Doctor Name"]
    data = [[r[2] or "", r[1], r[3] or "", r[4] or "", dnames.get(r[5], "")] for r in rows]
    return data, headers, len(rows)


# ═══════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR  –  professional formatting
# ═══════════════════════════════════════════════════════════════════
def _generate_excel(data, headers, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    even_fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = even_fill
            if isinstance(val, float) and col_idx > 1:
                has_currency = any(kw in (headers[col_idx - 1] if col_idx - 1 < len(headers) else "") for kw in ["Amount", "Total", "Paid", "Pending", "Bill", "Revenue", "Expense"])
                if has_currency:
                    cell.number_format = '#,##0.00'
    ws.freeze_panes = "A2"
    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1])) + 2
        for row_idx in range(2, min(len(data) + 2, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)) + 2, 50))
        ws.column_dimensions[col_letter].width = max_len
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════════
#  PDF GENERATOR  –  professional, no technical fields
# ═══════════════════════════════════════════════════════════════════
def _add_fonts(pdf):
    """Register Unicode-capable Arial font from Windows."""
    font_dir = "C:/Windows/Fonts"
    pdf.add_font("Arial", "", f"{font_dir}/arial.ttf", uni=True)
    pdf.add_font("Arial", "B", f"{font_dir}/arialbd.ttf", uni=True)
    pdf.add_font("Arial", "I", f"{font_dir}/ariali.ttf", uni=True)
    pdf.add_font("Arial", "BI", f"{font_dir}/arialbi.ttf", uni=True)

def _pdf_header(pdf, title, info, date_start, date_end):
    from fpdf import FPDF
    _add_fonts(pdf)
    pdf.set_fill_color(44, 62, 80)
    pdf.rect(0, 0, 210, 38 if info.get("name") else 30, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 18)
    pdf.set_y(8)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    if info.get("name"):
        pdf.set_font("Arial", "", 8)
        pdf.cell(0, 5, info["name"], new_x="LMARGIN", new_y="NEXT", align="C")
    if info.get("address"):
        pdf.set_font("Arial", "", 7)
        pdf.cell(0, 4, info["address"], new_x="LMARGIN", new_y="NEXT", align="C")
    if info.get("phone"):
        pdf.set_font("Arial", "", 7)
        pdf.cell(0, 4, f"Phone: {info['phone']}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(3)
    pdf.set_text_color(100, 100, 100)
    pdf.set_font("Arial", "", 7)
    period_str = f"Period: {date_start.strftime('%d-%b-%Y')} to {(date_end - timedelta(days=1)).strftime('%d-%b-%Y')}" if date_start and date_end else ""
    gen_str = f"Generated: {datetime.now(timezone.utc).strftime('%d-%b-%Y %H:%M')}"
    pdf.cell(95, 4, period_str, align="L")
    pdf.cell(95, 4, gen_str, align="R")
    pdf.ln(6)


def _pdf_table(pdf, headers, data, col_widths=None):
    from fpdf import FPDF
    effective_w = 190
    if col_widths:
        cw = col_widths
    else:
        cw = [effective_w / len(headers)] * len(headers)
    pdf.set_font("Arial", "B", 7)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(cw[i], 7, h[:25], border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Arial", "", 7)
    for row_idx, row in enumerate(data):
        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.set_font("Arial", "B", 7)
            pdf.set_fill_color(44, 62, 80)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(cw[i], 7, h[:25], border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_text_color(30, 30, 30)
            pdf.set_font("Arial", "", 7)
        if row_idx % 2 == 1:
            pdf.set_fill_color(245, 245, 245)
            fill = True
        else:
            fill = False
        for i, val in enumerate(row):
            txt = str(val)[:30] if val is not None else ""
            pdf.cell(cw[i], 6, txt, border=1, fill=fill, align="C" if i == 0 else "L")
        pdf.ln()


def _pdf_footer(pdf):
    pdf.set_y(-15)
    pdf.set_font("Arial", "I", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 10, f"Page {pdf.page_no()}/{{nb}}", align="C")


async def _generate_pdf(title, headers, data, filename, info=None, date_start=None, date_end=None):
    from fpdf import FPDF
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.add_page()
    _pdf_header(pdf, title, info or {}, date_start, date_end)
    _pdf_table(pdf, headers, data)
    _pdf_footer(pdf)
    filepath = os.path.join(EXPORT_DIR, filename)
    pdf.output(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════════
#  CHART GENERATORS  (matplotlib)
# ═══════════════════════════════════════════════════════════════════
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


def _save_chart(fig, filename):
    filepath = os.path.join(EXPORT_DIR, filename)
    fig.savefig(filepath, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return filepath


async def _revenue_trend_chart(db, hospital_ids, date_start, date_end, filename):
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)
    range_days = (date_end - date_start).days
    pg_fmt = "YYYY-MM-DD" if range_days <= 31 else "YYYY-MM"
    sq_fmt = "%Y-%m-%d" if range_days <= 31 else "%Y-%m"
    period_expr = func.to_char(Billing.updated_at, pg_fmt) if settings.DB_IS_POSTGRESQL else func.strftime(sq_fmt, Billing.updated_at)
    q = select(
        period_expr.label("period"),
        func.sum(Billing.paid_amount).label("revenue"),
    ).where(Billing.updated_at >= date_start, Billing.updated_at < date_end)
    if cids is not None:
        q = q.where(Billing.case_id.in_(cids))
    q = q.group_by("period").order_by("period")
    rows = (await db.execute(q)).all()
    periods = [r[0] for r in rows]
    revenues = [float(r[1]) for r in rows]
    fig, ax = plt.subplots(figsize=(8, 3.2))
    ax.plot(periods, revenues, marker="o", color="#2C3E50", linewidth=2)
    ax.fill_between(periods, revenues, alpha=0.1, color="#2C3E50")
    ax.set_title("Revenue Trend", fontsize=13, fontweight="bold", color="#2C3E50")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"\u20B9{x:,.0f}"))
    fig.tight_layout()
    return _save_chart(fig, filename)


async def _appointment_trend_chart(db, hospital_ids, date_start, date_end, filename):
    pids = await _get_patient_ids(db, hospital_ids)
    range_days = (date_end - date_start).days
    pg_fmt = "YYYY-MM-DD" if range_days <= 31 else "YYYY-MM"
    sq_fmt = "%Y-%m-%d" if range_days <= 31 else "%Y-%m"
    period_expr = func.to_char(Appointment.appointment_date, pg_fmt) if settings.DB_IS_POSTGRESQL else func.strftime(sq_fmt, Appointment.appointment_date)
    q = select(
        period_expr.label("period"),
        func.count(Appointment.id).label("count"),
    ).where(Appointment.appointment_date >= date_start.date() if hasattr(date_start, 'date') else date_start,
             Appointment.appointment_date < date_end.date() if hasattr(date_end, 'date') else date_end)
    if pids is not None:
        q = q.where(Appointment.patient_id.in_(pids))
    q = q.group_by("period").order_by("period")
    rows = (await db.execute(q)).all()
    periods = [r[0] for r in rows]
    counts = [int(r[1]) for r in rows]
    fig, ax = plt.subplots(figsize=(8, 3.2))
    ax.bar(periods, counts, color="#3498DB", alpha=0.8)
    ax.set_title("Appointment Trend", fontsize=13, fontweight="bold", color="#2C3E50")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    fig.tight_layout()
    return _save_chart(fig, filename)


async def _patient_growth_chart(db, hospital_ids, date_start, date_end, filename):
    range_days = (date_end - date_start).days
    pg_fmt = "YYYY-MM-DD" if range_days <= 31 else "YYYY-MM"
    sq_fmt = "%Y-%m-%d" if range_days <= 31 else "%Y-%m"
    period_expr = func.to_char(Patient.created_at, pg_fmt) if settings.DB_IS_POSTGRESQL else func.strftime(sq_fmt, Patient.created_at)
    q = select(
        period_expr.label("period"),
        func.count(Patient.id).label("count"),
    ).where(Patient.created_at >= date_start, Patient.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(Patient.hospital_id.in_(hospital_ids))
    q = q.group_by("period").order_by("period")
    rows = (await db.execute(q)).all()
    counts = [int(r[1]) for r in rows]
    cumulative = []
    s = 0
    for c in counts:
        s += c
        cumulative.append(s)
    periods = [r[0] for r in rows]
    fig, ax = plt.subplots(figsize=(8, 3.2))
    ax.fill_between(periods, cumulative, alpha=0.3, color="#27AE60")
    ax.plot(periods, cumulative, color="#27AE60", linewidth=2, marker="o")
    ax.set_title("Patient Growth", fontsize=13, fontweight="bold", color="#2C3E50")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    fig.tight_layout()
    return _save_chart(fig, filename)


async def _lead_source_chart(db, hospital_ids, date_start, date_end, filename):
    q = select(
        Lead.source, func.count(Lead.id).label("count"),
    ).where(Lead.created_at >= date_start, Lead.created_at < date_end)
    if hospital_ids is not None:
        q = q.where(Lead.hospital_id.in_(hospital_ids))
    q = q.group_by(Lead.source).order_by(func.count(Lead.id).desc())
    rows = (await db.execute(q)).all()
    labels = [r[0] or "Unknown" for r in rows]
    values = [int(r[1]) for r in rows]
    colors = ["#3498DB", "#2ECC71", "#E74C3C", "#F39C12", "#9B59B6", "#1ABC9C", "#E67E22"]
    fig, ax = plt.subplots(figsize=(5, 3.5))
    wedges, texts, autotexts = ax.pie(
        values, labels=None, autopct="%1.0f%%", startangle=90,
        colors=colors[:len(values)], pctdistance=0.75,
    )
    ax.set_title("Lead Sources", fontsize=13, fontweight="bold", color="#2C3E50")
    ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(1, 0.5), fontsize=7)
    fig.tight_layout()
    return _save_chart(fig, filename)


async def _doctor_revenue_chart(db, hospital_ids, date_start, date_end, filename):
    q = select(
        Case.doctor_id, func.sum(Billing.paid_amount).label("revenue"),
    ).join(Billing, Billing.case_id == Case.id
    ).where(Billing.updated_at >= date_start, Billing.updated_at < date_end)
    if hospital_ids is not None:
        pids = await _get_patient_ids(db, hospital_ids)
        cids = await _get_case_ids(db, pids)
        if cids is not None:
            q = q.where(Case.id.in_(cids))
    q = q.group_by(Case.doctor_id).order_by(func.sum(Billing.paid_amount).desc())
    rows = (await db.execute(q)).all()
    doc_names = []
    revenues = []
    for did, rev in rows:
        if did:
            dr = await db.execute(select(User.full_name).where(User.id == did))
            doc_names.append(dr.scalar() or "Unknown")
        else:
            doc_names.append("Unassigned")
        revenues.append(float(rev))
    fig, ax = plt.subplots(figsize=(7, 3.5))
    colors_bar = ["#2C3E50", "#3498DB", "#2ECC71", "#F39C12", "#E74C3C", "#9B59B6"]
    ax.barh(doc_names, revenues, color=colors_bar[:len(doc_names)])
    ax.set_title("Doctor Revenue", fontsize=13, fontweight="bold", color="#2C3E50")
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"\u20B9{x:,.0f}"))
    fig.tight_layout()
    return _save_chart(fig, filename)


async def _treatment_performance_chart(db, hospital_ids, date_start, date_end, filename):
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)
    q = select(
        TreatmentPlan.treatment_name, func.count(TreatmentPlan.id).label("count"),
    ).where(TreatmentPlan.created_at >= date_start, TreatmentPlan.created_at < date_end)
    if cids is not None:
        q = q.where(TreatmentPlan.case_id.in_(cids))
    q = q.group_by(TreatmentPlan.treatment_name).order_by(func.count(TreatmentPlan.id).desc()).limit(10)
    rows = (await db.execute(q)).all()
    names = [r[0] or "Unknown" for r in rows]
    counts = [int(r[1]) for r in rows]
    fig, ax = plt.subplots(figsize=(7, 3.5))
    colors_bar = ["#2C3E50", "#3498DB", "#2ECC71", "#F39C12", "#E74C3C", "#9B59B6", "#1ABC9C", "#E67E22", "#2980B9", "#8E44AD"]
    ax.barh(names, counts, color=colors_bar[:len(names)])
    ax.set_title("Top Treatments", fontsize=13, fontweight="bold", color="#2C3E50")
    fig.tight_layout()
    return _save_chart(fig, filename)


# ═══════════════════════════════════════════════════════════════════
#  DASHBOARD PDF  –  executive summary + charts
# ═══════════════════════════════════════════════════════════════════
async def generate_dashboard_pdf(
    db: AsyncSession, current_user: dict,
    period: str = "this_month", start_date: str = None, end_date: str = None,
) -> str:
    from fpdf import FPDF
    hospital_ids = await _get_hospital_ids(db, current_user)
    date_start, date_end = get_date_range(period, start_date, end_date)
    user_name = current_user.get("full_name", "User")
    now_str = datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M")
    hid = current_user.get("hospital_id")
    info = await _hospital_info(db, hid)
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)

    # Generate charts
    chart_files = []
    try:
        chart_files.append(await _revenue_trend_chart(db, hospital_ids, date_start, date_end, f"ch_rev_{uuid.uuid4().hex[:8]}.png"))
        chart_files.append(await _appointment_trend_chart(db, hospital_ids, date_start, date_end, f"ch_appt_{uuid.uuid4().hex[:8]}.png"))
        chart_files.append(await _patient_growth_chart(db, hospital_ids, date_start, date_end, f"ch_pat_{uuid.uuid4().hex[:8]}.png"))
        chart_files.append(await _lead_source_chart(db, hospital_ids, date_start, date_end, f"ch_lead_{uuid.uuid4().hex[:8]}.png"))
        chart_files.append(await _doctor_revenue_chart(db, hospital_ids, date_start, date_end, f"ch_doc_{uuid.uuid4().hex[:8]}.png"))
        chart_files.append(await _treatment_performance_chart(db, hospital_ids, date_start, date_end, f"ch_treat_{uuid.uuid4().hex[:8]}.png"))
    except Exception:
        pass

    # Metrics
    r = await db.execute(select(func.sum(Billing.paid_amount)).where(
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    total_revenue = float(r.scalar() or 0)

    r = await db.execute(select(func.count(Patient.id)).where(
        Patient.created_at >= date_start, Patient.created_at < date_end,
        *((Patient.hospital_id.in_(hospital_ids)) if hospital_ids is not None else ())
    ))
    new_patients = int(r.scalar() or 0)

    r = await db.execute(select(func.count(Appointment.id)).where(
        Appointment.appointment_date >= date_start.date() if hasattr(date_start, 'date') else date_start,
        Appointment.appointment_date < date_end.date() if hasattr(date_end, 'date') else date_end,
        *((Appointment.patient_id.in_(pids)) if pids is not None else ())
    ))
    total_appointments = int(r.scalar() or 0)

    r = await db.execute(select(func.count(TreatmentPlan.id)).where(
        TreatmentPlan.created_at >= date_start, TreatmentPlan.created_at < date_end,
        *(TreatmentPlan.case_id.in_(cids) if cids is not None else ())
    ))
    total_treatments = int(r.scalar() or 0)

    r = await db.execute(select(func.count(Billing.id)).where(
        Billing.payment_status.in_(["PAID", "PARTIAL"]),
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    collected_count = int(r.scalar() or 0)
    r = await db.execute(select(func.sum(Billing.paid_amount)).where(
        Billing.payment_status.in_(["PAID", "PARTIAL"]),
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    collected_amount = float(r.scalar() or 0)

    r = await db.execute(select(func.sum(Billing.pending_amount)).where(
        Billing.payment_status.in_(["PARTIAL", "OVERDUE"]),
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    outstanding = float(r.scalar() or 0)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.add_page()
    _pdf_header(pdf, "Dashboard Report", info, date_start, date_end)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(44, 62, 80)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    for label, val in [
        ("Total Revenue", f"\u20B9{total_revenue:,.2f}"),
        ("New Patients", str(new_patients)),
        ("Appointments", str(total_appointments)),
        ("Treatments Started", str(total_treatments)),
        ("Collections", f"\u20B9{collected_amount:,.2f} ({collected_count} invoices)"),
        ("Outstanding", f"\u20B9{outstanding:,.2f}"),
    ]:
        pdf.set_font("Arial", "B", 9)
        pdf.cell(50, 6, label)
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Charts & Analytics", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)
    for i, cp in enumerate(chart_files):
        try:
            if i > 0:
                pdf.add_page()
            pdf.image(cp, x=10, w=190)
            pdf.ln(2)
        except Exception:
            pass
    _pdf_footer(pdf)
    filepath = os.path.join(EXPORT_DIR, f"dashboard_report_{datetime.now(timezone.utc).strftime('%Y_%m')}.pdf")
    pdf.output(filepath)
    for cp in chart_files:
        try:
            os.remove(cp)
        except Exception:
            pass
    return filepath


# ═══════════════════════════════════════════════════════════════════
#  FINANCIAL REPORT PDF
# ═══════════════════════════════════════════════════════════════════
async def generate_financial_report_pdf(
    db: AsyncSession, current_user: dict,
    period: str = "this_month", start_date: str = None, end_date: str = None,
) -> str:
    from fpdf import FPDF
    hospital_ids = await _get_hospital_ids(db, current_user)
    date_start, date_end = get_date_range(period, start_date, end_date)
    user_name = current_user.get("full_name", "User")
    now_str = datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M")
    hid = current_user.get("hospital_id")
    info = await _hospital_info(db, hid)
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)

    r = await db.execute(select(func.sum(Billing.total_amount)).where(
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    total_billed = float(r.scalar() or 0)

    r = await db.execute(select(func.sum(Billing.paid_amount)).where(
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    total_collected = float(r.scalar() or 0)

    r = await db.execute(select(func.sum(Billing.pending_amount)).where(
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    total_outstanding = float(r.scalar() or 0)

    total_expenses = await calculate_expenses_for_date_range(db, hospital_ids, date_start=date_start, date_end=date_end)
    net_profit = total_collected - total_expenses
    profit_margin = round((net_profit / total_collected * 100), 2) if total_collected > 0 else 0

    r = await db.execute(select(
        Case.doctor_id, func.sum(Billing.paid_amount).label("rev"),
    ).join(Billing, Billing.case_id == Case.id
    ).where(Billing.updated_at >= date_start, Billing.updated_at < date_end
    ).group_by(Case.doctor_id).order_by(func.sum(Billing.paid_amount).desc()))
    doc_rows = r.all()

    r = await db.execute(select(
        TreatmentPlan.treatment_name, func.sum(Billing.paid_amount).label("rev"),
    ).join(Billing, Billing.case_id == TreatmentPlan.case_id
    ).where(Billing.updated_at >= date_start, Billing.updated_at < date_end
    ).group_by(TreatmentPlan.treatment_name).order_by(func.sum(Billing.paid_amount).desc()).limit(10))
    treat_rows = r.all()

    r = await db.execute(select(
        Patient.patient_source, func.sum(Billing.paid_amount).label("rev"),
    ).join(Case, Case.patient_id == Patient.id
    ).join(Billing, Billing.case_id == Case.id
    ).where(Billing.updated_at >= date_start, Billing.updated_at < date_end,
            Patient.patient_source.isnot(None)
    ).group_by(Patient.patient_source).order_by(func.sum(Billing.paid_amount).desc()))
    source_rows = r.all()

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.add_page()
    _pdf_header(pdf, "Financial Report", info, date_start, date_end)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Revenue Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(44, 62, 80)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    for label, val in [
        ("Total Billed", f"\u20B9{total_billed:,.2f}"),
        ("Total Collected", f"\u20B9{total_collected:,.2f}"),
        ("Outstanding", f"\u20B9{total_outstanding:,.2f}"),
        ("Total Expenses", f"\u20B9{total_expenses:,.2f}"),
        ("Net Profit", f"\u20B9{net_profit:,.2f}"),
        ("Profit Margin", f"{profit_margin}%"),
    ]:
        pdf.set_font("Arial", "B", 9)
        pdf.cell(50, 6, label)
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Revenue by Doctor", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)
    pdf.set_text_color(0, 0, 0)
    for did, rev in doc_rows:
        if did:
            dr = await db.execute(select(User.full_name).where(User.id == did))
            dname = dr.scalar() or "Unknown"
        else:
            dname = "Unassigned"
        pdf.set_font("Arial", "", 9)
        pdf.cell(80, 6, dname)
        pdf.cell(0, 6, f"\u20B9{float(rev):,.2f}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Revenue by Treatment", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)
    pdf.set_text_color(0, 0, 0)
    for tname, rev in treat_rows:
        pdf.set_font("Arial", "", 9)
        pdf.cell(80, 6, tname or "Unknown")
        pdf.cell(0, 6, f"\u20B9{float(rev):,.2f}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Revenue by Source", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)
    pdf.set_text_color(0, 0, 0)
    for sname, rev in source_rows:
        pdf.set_font("Arial", "", 9)
        pdf.cell(80, 6, sname or "Unknown")
        pdf.cell(0, 6, f"\u20B9{float(rev):,.2f}", new_x="LMARGIN", new_y="NEXT")
    _pdf_footer(pdf)
    filepath = os.path.join(EXPORT_DIR, f"financial_report_{datetime.now(timezone.utc).strftime('%Y_%m')}.pdf")
    pdf.output(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════════
#  MONTHLY MANAGEMENT REPORT PDF
# ═══════════════════════════════════════════════════════════════════
async def generate_monthly_report_pdf(
    db: AsyncSession, current_user: dict,
    period: str = "this_month", start_date: str = None, end_date: str = None,
) -> str:
    from fpdf import FPDF
    hospital_ids = await _get_hospital_ids(db, current_user)
    date_start, date_end = get_date_range(period, start_date, end_date)
    user_name = current_user.get("full_name", "User")
    now_str = datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M")
    report_month = date_start.strftime("%B %Y")
    hid = current_user.get("hospital_id")
    info = await _hospital_info(db, hid)
    pids = await _get_patient_ids(db, hospital_ids)
    cids = await _get_case_ids(db, pids)

    r = await db.execute(select(func.sum(Billing.paid_amount)).where(
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    total_revenue = float(r.scalar() or 0)

    r = await db.execute(select(func.count(Patient.id)).where(
        Patient.created_at >= date_start, Patient.created_at < date_end,
        *((Patient.hospital_id.in_(hospital_ids)) if hospital_ids is not None else ())
    ))
    new_patients = int(r.scalar() or 0)

    r = await db.execute(select(func.count(Appointment.id)).where(
        Appointment.appointment_date >= date_start.date() if hasattr(date_start, 'date') else date_start,
        Appointment.appointment_date < date_end.date() if hasattr(date_end, 'date') else date_end,
        *((Appointment.patient_id.in_(pids)) if pids is not None else ())
    ))
    total_appointments = int(r.scalar() or 0)

    r = await db.execute(select(func.count(TreatmentPlan.id)).where(
        TreatmentPlan.created_at >= date_start, TreatmentPlan.created_at < date_end,
        *(TreatmentPlan.case_id.in_(cids) if cids is not None else ())
    ))
    total_treatments = int(r.scalar() or 0)

    total_expenses = await calculate_expenses_for_date_range(db, hospital_ids, date_start=date_start, date_end=date_end)
    net_profit = total_revenue - total_expenses
    pm = round((net_profit / total_revenue * 100), 2) if total_revenue > 0 else 0

    r = await db.execute(select(func.count(Lead.id)).where(
        Lead.created_at >= date_start, Lead.created_at < date_end,
        *((Lead.hospital_id.in_(hospital_ids)) if hospital_ids is not None else ())
    ))
    total_leads = int(r.scalar() or 0)

    r = await db.execute(select(func.count(FollowUp.id)).where(
        FollowUp.created_at >= date_start, FollowUp.created_at < date_end,
        *((FollowUp.hospital_id.in_(hospital_ids)) if hospital_ids is not None else ())
    ))
    total_followups = int(r.scalar() or 0)

    recall_types = ("6_MONTH_RECALL", "12_MONTH_RECALL", "CUSTOM_RECALL", "CUSTOM_FOLLOW_UP")
    r = await db.execute(select(func.count(FollowUp.id)).where(
        FollowUp.follow_up_type.in_(recall_types),
        FollowUp.created_at >= date_start, FollowUp.created_at < date_end,
        *((FollowUp.hospital_id.in_(hospital_ids)) if hospital_ids is not None else ())
    ))
    total_recalls = int(r.scalar() or 0)

    r = await db.execute(select(func.sum(Billing.pending_amount)).where(
        Billing.payment_status.in_(["PARTIAL", "OVERDUE"]),
        Billing.updated_at >= date_start, Billing.updated_at < date_end,
        *(Billing.case_id.in_(cids) if cids is not None else ())
    ))
    outstanding = float(r.scalar() or 0)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    _add_fonts(pdf)
    pdf.add_page()
    pdf.set_fill_color(44, 62, 80)
    pdf.rect(0, 0, 210, 50, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 20)
    pdf.set_y(10)
    pdf.cell(0, 12, "Monthly Management Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 7, report_month, new_x="LMARGIN", new_y="NEXT", align="C")
    if info.get("name"):
        pdf.set_font("Arial", "", 8)
        pdf.cell(0, 5, info["name"], new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Arial", "", 7)
    pdf.cell(0, 5, f"Generated: {now_str}  |  By: {user_name}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(8)
    pdf.set_text_color(44, 62, 80)
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 8, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(44, 62, 80)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    for label, val in [
        ("Revenue", f"\u20B9{total_revenue:,.2f}"),
        ("Expenses", f"\u20B9{total_expenses:,.2f}"),
        ("Net Profit", f"\u20B9{net_profit:,.2f}"),
        ("Profit Margin", f"{pm}%"),
        ("Outstanding", f"\u20B9{outstanding:,.2f}"),
        ("New Patients", str(new_patients)),
        ("Appointments", str(total_appointments)),
        ("Treatments", str(total_treatments)),
        ("New Leads", str(total_leads)),
        ("Follow-ups", str(total_followups)),
        ("Recalls", str(total_recalls)),
    ]:
        pdf.set_font("Arial", "B", 9)
        pdf.cell(50, 6, label)
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")
    _pdf_footer(pdf)
    filepath = os.path.join(EXPORT_DIR, f"monthly_report_{datetime.now(timezone.utc).strftime('%Y_%m')}.pdf")
    pdf.output(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════════
#  MAIN EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════════════
EXPORT_MODULES = {
    "patients": _csv_patients,
    "appointments": _csv_appointments,
    "cases": _csv_cases,
    "treatments": _csv_treatments,
    "billings": _csv_billings,
    "expenses": _csv_expenses,
    "leads": _csv_leads,
    "enquiries": _csv_enquiries,
    "follow-ups": _csv_follow_ups,
    "recalls": _csv_recalls,
    "doctors": _csv_doctors,
    "consent-forms": _csv_consent_forms,
}

MODULE_LABELS = {
    "patients": "Patients", "appointments": "Appointments", "cases": "Cases",
    "treatments": "Treatments", "billings": "Billings", "expenses": "Expenses",
    "leads": "Leads", "enquiries": "Enquiries", "follow-ups": "Follow-Ups",
    "recalls": "Recalls", "doctors": "Doctors", "consent-forms": "Consent Forms",
    "doctor-performance": "Doctor Performance",
}


async def export_data(
    db: AsyncSession, current_user: dict, module: str, format: str,
    period: str = "this_month", start_date: str = None, end_date: str = None,
):
    hospital_ids = await _get_hospital_ids(db, current_user)
    date_start, date_end = get_date_range(period, start_date, end_date)
    if module == "doctor-performance":
        data, headers, count = await _csv_doctor_performance(db, current_user, date_start, date_end)
    else:
        generator = EXPORT_MODULES.get(module)
        if not generator:
            raise ValueError(f"Unknown module: {module}")
        data, headers, count = await generator(db, hospital_ids, date_start, date_end)
    label = MODULE_LABELS.get(module, module)
    date_str = datetime.now(timezone.utc).strftime("%Y_%m_%d")
    hid = current_user.get("hospital_id")
    info = await _hospital_info(db, hid)

    if format == "csv":
        filename = f"{label.lower()}_{date_str}.csv"
        return _stream_csv(data, headers, filename), count

    elif format == "excel":
        safe = label.lower().replace(" ", "_")
        filename = f"{safe}_{date_str}.xlsx"
        # openpyxl rendering is pure CPU and can be large — keep it off the
        # event loop so concurrent requests are not starved.
        import asyncio
        filepath = await asyncio.to_thread(_generate_excel, data, headers, filename)
        return filepath, count

    elif format == "pdf":
        safe = label.lower().replace(" ", "_")
        filename = f"{safe}_{date_str}.pdf"
        filepath = await _generate_pdf(label, headers, data, filename, info=info, date_start=date_start, date_end=date_end)
        return filepath, count

    else:
        raise ValueError(f"Unknown format: {format}")


async def log_export(db: AsyncSession, current_user: dict, module: str, format: str, record_count: int, file_path: str = None):
    log = AuditLog(
        user_id=current_user.get("sub"),
        action="EXPORT",
        entity_type=f"export_{module}",
        entity_id=file_path,
        details=f"Exported {record_count} {module} records as {format}",
        ip_address=None,
    )
    db.add(log)
    await db.commit()



